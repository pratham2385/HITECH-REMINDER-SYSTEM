import pandas as pd
import openpyxl
from openpyxl import Workbook
import os, sys

sys.path.append('C:/Users/admin/Downloads/HITECH-REMINDER-SYSTEM')
from src.services.excel_importer import _load_workbook, _sheet_rows, _find_header_row, _make_headers

wb = Workbook()
ws = wb.active
ws.append(["Col1", "Col2", "Col3", "Col4", "Col5", "Col6", "Col7"])
ws.append(["A", "B", "C", "D", "E", "F", "G"])
wb.save('scratch/test.xlsx')

wb2 = _load_workbook('scratch/test.xlsx')
ws2 = wb2.active
rows = _sheet_rows(ws2)
print("Row 0 length:", len(rows[0]))
print("Column count:", max((len(r) for r in rows), default=0))

header_idx = _find_header_row(rows)
print("Header index:", header_idx)
print("Headers:", _make_headers(rows[header_idx], max((len(r) for r in rows), default=0)))
