"""Excel import service for turning workbooks into dashboard modules."""

from __future__ import annotations

import json
import shutil
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from uuid import uuid4

from sqlalchemy.orm import Session

from src.db.models import (
    ActivityRecord,
    ImportedSheet,
    Module,
    ModuleDataRecord,
    ModuleField,
    WorkbookImport,
    User,
    NotificationSetting,
    TaskCollection,
)
from src.utils.helpers import clean_string


@dataclass(frozen=True, slots=True)
class SheetPreview:
    """Preview metadata for one workbook sheet."""

    name: str
    columns: list[str]
    row_count: int
    sample_rows: list[dict[str, Any]]
    is_activity_like: bool


@dataclass(frozen=True, slots=True)
class WorkbookPreview:
    """Preview metadata for one workbook."""

    filename: str
    sheet_count: int
    row_count: int
    sheets: list[SheetPreview]


def _load_workbook(path: Path):
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("Missing required dependency: openpyxl. Run `pip install -r requirements.txt`.") from exc

    return load_workbook(path, data_only=True, read_only=False)


def _json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return value


def _cell_value(cell: Any) -> Any:
    value = _json_value(cell.value)
    if cell.hyperlink and cell.hyperlink.target:
        return {"value": value if value is not None else cell.hyperlink.target, "link": cell.hyperlink.target}
    return value


def _display_value(value: Any) -> str:
    if isinstance(value, dict):
        return clean_string(value.get("value", ""))
    return clean_string(value)


def _make_headers(raw_headers: list[Any], column_count: int) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}

    for index in range(column_count):
        base = _display_value(raw_headers[index]) if index < len(raw_headers) else ""
        header = base or f"Column {index + 1}"
        if header in seen:
            seen[header] += 1
            header = f"{header} {seen[header]}"
        else:
            seen[header] = 1
        headers.append(header)

    return headers


def _find_header_row(rows: list[list[Any]]) -> int:
    for index, row in enumerate(rows):
        if any(_display_value(value) for value in row):
            return index
    return 0


def _sheet_rows(sheet: Any) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in sheet.iter_rows():
        rows.append([_cell_value(cell) for cell in row])
    return rows


def _row_dict(columns: list[str], row: list[Any]) -> dict[str, Any]:
    return {column: row[index] if index < len(row) else None for index, column in enumerate(columns)}


def _is_activity_like(columns: list[str]) -> bool:
    normalized = {column.strip().casefold() for column in columns}
    return {"activity", "frequency", "date"}.issubset(normalized)


class ExcelImportService:
    """Imports uploaded Excel workbooks into modules and activities."""

    def __init__(self, upload_dir: Path) -> None:
        self.upload_dir = upload_dir

    def save_uploaded_file(self, filename: str, source_path: Path) -> Path:
        """Copy an uploaded workbook into the managed upload directory."""

        self.upload_dir.mkdir(parents=True, exist_ok=True)
        safe_name = Path(filename).name or "workbook.xlsx"
        stored_path = self.upload_dir / f"{uuid4().hex}_{safe_name}"
        shutil.copy2(source_path, stored_path)
        return stored_path

    def preview_workbook(self, workbook_path: Path) -> WorkbookPreview:
        """Return a structured preview of a workbook without importing it."""

        workbook = _load_workbook(workbook_path)
        previews: list[SheetPreview] = []
        total_rows = 0

        for sheet in workbook.worksheets:
            raw_rows = _sheet_rows(sheet)
            if not raw_rows:
                previews.append(SheetPreview(sheet.title, [], 0, [], False))
                continue

            column_count = max((len(row) for row in raw_rows), default=0)
            header_index = _find_header_row(raw_rows)
            columns = _make_headers(raw_rows[header_index] if raw_rows else [], column_count)
            data_rows = raw_rows[header_index + 1 :]
            row_count = sum(1 for row in data_rows if any(_display_value(value) for value in row))
            total_rows += row_count
            sample_rows = [_row_dict(columns, row) for row in data_rows[:50]]

            previews.append(
                SheetPreview(
                    name=sheet.title,
                    columns=columns,
                    row_count=row_count,
                    sample_rows=sample_rows,
                    is_activity_like=_is_activity_like(columns),
                )
            )

        return WorkbookPreview(
            filename=workbook_path.name,
            sheet_count=len(previews),
            row_count=total_rows,
            sheets=previews,
        )

    def create_pending_import(
        self,
        session: Session,
        original_filename: str,
        stored_path: Path,
        imported_by_user_id: int | None,
    ) -> WorkbookImport:
        """Create a pending import history row after upload."""

        preview = self.preview_workbook(stored_path)
        record = WorkbookImport(
            original_filename=original_filename,
            stored_path=str(stored_path),
            status="pending",
            sheet_count=preview.sheet_count,
            row_count=preview.row_count,
            imported_by_user_id=imported_by_user_id,
        )
        session.add(record)
        session.flush()
        return record

    
    def import_workbook(
        self,
        session: Session,
        workbook_import: WorkbookImport,
        import_activity_sheets: bool = True,
    ) -> WorkbookImport:
        """Import workbook based on the flattened structure."""

        workbook_path = Path(workbook_import.stored_path)
        workbook = _load_workbook(workbook_path)
        total_rows = 0

        # Deactivate all existing workspaces
        for c in session.query(TaskCollection).all():
            c.is_active = False

        # Create Workspace (TaskCollection) for this import
        workspace = TaskCollection(
            name=f"Import: {workbook_import.original_filename}",
            description=f"Auto-generated workspace for import {workbook_import.id}",
            import_date=datetime.utcnow(),
            imported_file_name=workbook_import.original_filename,
            is_active=True
        )
        session.add(workspace)
        session.flush()

        from src.db.models import Recipient

        # Only process the first sheet (or assume the primary data is on the first sheet)
        if workbook.worksheets:
            sheet = workbook.worksheets[0]
            raw_rows = _sheet_rows(sheet)
            header_index = _find_header_row(raw_rows) if raw_rows else 0
            
            if raw_rows:
                column_count = max((len(row) for row in raw_rows), default=0)
                columns = _make_headers(raw_rows[header_index] if raw_rows else [], column_count)
                data_rows = raw_rows[header_index + 1 :] if raw_rows else []
                nonblank_rows = [row for row in data_rows if any(_display_value(value) for value in row)]

                column_lookup = {col.strip().casefold(): col for col in columns}
                def get_value(row_dict, key): 
                    val = row_dict.get(column_lookup.get(key), "")
                    if isinstance(val, dict):
                        return clean_string(val.get("value", ""))
                    return clean_string(val)

                for index, row in enumerate(nonblank_rows, start=1):
                    row_dict = _row_dict(columns, row)
                    
                    # 1. Module -> Category
                    category_name = get_value(row_dict, "module")

                    # 2. Recipient
                    recipient_email = get_value(row_dict, "recipient_email")
                    recipient = None
                    if recipient_email:
                        recipient = session.query(Recipient).filter(
                            Recipient.email == recipient_email,
                            Recipient.workspace_id == workspace.id
                        ).first()
                        if not recipient:
                            recipient = Recipient(
                                email=recipient_email, 
                                name=get_value(row_dict, "recipient_name") or recipient_email,
                                workspace_id=workspace.id
                            )
                            session.add(recipient)
                            session.flush()

                    # 3. Activity (Fallback to common variations like Task Name)
                    activity_name = (
                        get_value(row_dict, "activity") or 
                        get_value(row_dict, "task_name") or 
                        get_value(row_dict, "task name") or 
                        get_value(row_dict, "task")
                    )
                    if not activity_name:
                        continue

                    # Parse Reminder Before Minutes safely
                    try:
                        reminder_mins = int(get_value(row_dict, "reminder_before_minutes") or 0)
                    except ValueError:
                        reminder_mins = 0

                    # Parse Start Date safely
                    start_date_str = get_value(row_dict, "start_date")
                    start_date = None
                    if start_date_str:
                        try:
                            from dateutil.parser import parse
                            start_date = parse(start_date_str)
                        except Exception:
                            pass
                            
                    # Parse Day of Month safely
                    try:
                        day_of_month = int(get_value(row_dict, "day_of_month") or 0)
                    except ValueError:
                        day_of_month = None

                    record = ActivityRecord(
                        activity=activity_name,
                        category=category_name,
                        frequency=get_value(row_dict, "frequency"),
                        send_time=get_value(row_dict, "time") or "09:00",
                        start_date=start_date,
                        day_of_week=get_value(row_dict, "day_of_week") or None,
                        day_of_month=day_of_month if day_of_month else None,
                        reminder_before_minutes=reminder_mins if reminder_mins else None,
                        linked_module_id=None,
                        recipient_id=recipient.id if recipient else None,
                        collection_id=workspace.id,
                        sort_order=index,
                        is_active=True,
                    )
                    session.add(record)

                total_rows = len(nonblank_rows)
                session.add(
                    ImportedSheet(
                        workbook_import_id=workbook_import.id,
                        module_id=None,
                        sheet_name=sheet.title,
                        row_count=total_rows,
                        column_count=len(columns),
                    )
                )

        workbook_import.status = "imported"
        workbook_import.sheet_count = 1
        workbook_import.row_count = total_rows
        workbook_import.completed_at = datetime.utcnow()
        session.flush()
        return workbook_import
