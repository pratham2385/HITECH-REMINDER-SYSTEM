import os
import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill
from pathlib import Path

def create_production_workbook(output_path: Path):
    wb = Workbook()
    
    # -------------------------------------------------------------------------
    # 1. Users Sheet
    # -------------------------------------------------------------------------
    ws_users = wb.active
    ws_users.title = "Users"
    ws_users.append(["User_ID", "Name", "Email", "Phone", "Role"])
    
    users_data = [
        ("USR-001", "Alice Accountant", "alice@example.com", "+1234567890", "Accountant"),
        ("USR-002", "Bob Bookkeeper", "bob@example.com", "+1987654321", "Accountant"),
        ("USR-003", "Charlie CFO", "charlie@example.com", "+1122334455", "Finance Manager"),
        ("USR-004", "Diana Director", "diana@example.com", "+1555666777", "Finance Manager"),
        ("USR-005", "Eve Executive", "eve@example.com", "+1999888777", "HR Executive"),
        ("USR-006", "Frank HR", "frank@example.com", "+1444333222", "HR Executive"),
        ("USR-007", "Grace Auditor", "grace@example.com", "+1777888999", "Auditor"),
        ("USR-008", "Hank Helper", "hank@example.com", "+1333222111", "Accountant"),
        ("USR-009", "Ivy Inspector", "ivy@example.com", "+1666555444", "Auditor"),
        ("USR-010", "Jack Junior", "jack@example.com", "+1222111000", "Accountant"),
        ("USR-011", "Karen Keys", "karen@example.com", "+1888777666", "Finance Manager"),
        ("USR-012", "Leo Ledger", "leo@example.com", "+1000999888", "Accountant"),
        ("USR-013", "Mona Money", "mona@example.com", "+1111222333", "Finance Manager"),
        ("USR-014", "Nina Numbers", "nina@example.com", "+1444555666", "Auditor"),
        ("USR-015", "Oscar Officer", "oscar@example.com", "+1777666555", "Compliance Officer")
    ]
    for row in users_data:
        ws_users.append(row)
        
    # -------------------------------------------------------------------------
    # 2. Modules Sheet
    # -------------------------------------------------------------------------
    ws_modules = wb.create_sheet(title="Modules")
    ws_modules.append(["Module_ID", "Module_Name", "Description"])
    
    modules_data = [
        ("MOD-001", "GST", "Goods and Services Tax filings and reconciliations."),
        ("MOD-002", "Payroll", "Salary, PF, ESIC, and PT processing."),
        ("MOD-003", "TDS", "Tax Deducted at Source payments and returns."),
        ("MOD-004", "Vendor Payments", "Accounts Payable and vendor reconciliations."),
        ("MOD-005", "Compliance", "Statutory and regulatory compliance."),
        ("MOD-006", "Audits", "Internal and external audit preparations."),
        ("MOD-007", "Financial Reports", "MIS, P&L, Balance Sheet preparations.")
    ]
    for row in modules_data:
        ws_modules.append(row)

    # -------------------------------------------------------------------------
    # 3. Activities Sheet
    # -------------------------------------------------------------------------
    ws_activities = wb.create_sheet(title="Activities")
    # Added new columns: Email_Enabled, WhatsApp_Enabled, Priority
    ws_activities.append(["Activity_ID", "Activity", "Frequency", "Date", "Module", "Assignee_ID", 
                          "Email_Enabled", "WhatsApp_Enabled", "Priority", "Status", "Link", "Remark"])
    
    activities_list = [
        # GST (MOD-001)
        ("ACT-001", "GSTR-1 Filing", "Monthly", "11", "GST", "USR-001", "Yes", "Yes", "High", "Pending", "https://gst.gov.in", "File outward supplies"),
        ("ACT-002", "GSTR-3B Filing", "Monthly", "20", "GST", "USR-001", "Yes", "Yes", "High", "Pending", "https://gst.gov.in", "Monthly return and payment"),
        ("ACT-003", "GSTR-2B Reconciliation", "Monthly", "14", "GST", "USR-008", "Yes", "No", "Medium", "In Progress", "", "Reconcile ITC"),
        ("ACT-004", "GSTR-9 Annual Return", "Yearly", "December", "GST", "USR-003", "Yes", "Yes", "High", "Pending", "", "Annual return filing"),
        ("ACT-005", "GSTR-9C Reconciliation", "Yearly", "December", "GST", "USR-007", "Yes", "Yes", "High", "Pending", "", "Audit certification"),
        ("ACT-006", "LUT Renewal", "Yearly", "March", "GST", "USR-015", "Yes", "No", "Medium", "Completed", "", "Letter of Undertaking for exports"),
        ("ACT-007", "E-Way Bill Reconciliation", "Weekly", "Friday", "GST", "USR-012", "Yes", "No", "Low", "Pending", "", "Check e-way bills generated vs sales"),
        
        # Payroll (MOD-002)
        ("ACT-008", "Salary Processing", "Monthly", "Last Day", "Payroll", "USR-005", "Yes", "Yes", "High", "Pending", "https://hr.system.com", "Finalize attendance and process salaries"),
        ("ACT-009", "PF Payment", "Monthly", "15", "Payroll", "USR-006", "Yes", "Yes", "High", "Pending", "https://epfindia.gov.in", "Provident Fund deposit"),
        ("ACT-010", "ESIC Payment", "Monthly", "15", "Payroll", "USR-006", "Yes", "Yes", "High", "Pending", "https://esic.in", "ESIC deposit"),
        ("ACT-011", "Professional Tax (PT) Payment", "Monthly", "20", "Payroll", "USR-005", "Yes", "No", "Medium", "Pending", "", "State-specific PT"),
        ("ACT-012", "TDS on Salary Payment", "Monthly", "7", "Payroll", "USR-005", "Yes", "Yes", "High", "Pending", "", "Deposit TDS deducted from salaries"),
        ("ACT-013", "Quarterly PF Return", "Quarterly", "15", "Payroll", "USR-006", "Yes", "No", "Medium", "Pending", "", "File PF returns"),
        ("ACT-014", "Annual Appraisals Review", "Yearly", "April", "Payroll", "USR-003", "Yes", "Yes", "High", "Pending", "", "Finalize increments"),

        # TDS (MOD-003)
        ("ACT-015", "TDS Payment (Non-Salary)", "Monthly", "7", "TDS", "USR-002", "Yes", "Yes", "High", "Pending", "https://incometax.gov.in", "Deposit TDS for contractors/rent/professionals"),
        ("ACT-016", "Form 26Q Filing", "Quarterly", "31", "TDS", "USR-002", "Yes", "Yes", "High", "Pending", "", "TDS return for non-salary"),
        ("ACT-017", "Form 24Q Filing", "Quarterly", "31", "TDS", "USR-005", "Yes", "Yes", "High", "Pending", "", "TDS return for salary"),
        ("ACT-018", "Form 27Q Filing", "Quarterly", "31", "TDS", "USR-002", "Yes", "No", "Medium", "Pending", "", "TDS return for NRI"),
        ("ACT-019", "Issue TDS Certificates (Form 16A)", "Quarterly", "15", "TDS", "USR-010", "Yes", "No", "Low", "Pending", "", "Distribute certificates to vendors"),
        ("ACT-020", "Issue Form 16 (Salary)", "Yearly", "June", "TDS", "USR-006", "Yes", "No", "Medium", "Pending", "", "Distribute Form 16 to employees"),

        # Vendor Payments (MOD-004)
        ("ACT-021", "Vendor Payment Run", "Weekly", "Wednesday", "Vendor Payments", "USR-004", "Yes", "No", "Medium", "Pending", "https://erp/ap", "Process approved vendor invoices"),
        ("ACT-022", "Vendor Ledger Reconciliation", "Monthly", "10", "Vendor Payments", "USR-012", "Yes", "No", "Low", "Pending", "", "Reconcile major vendors"),
        ("ACT-023", "MSME Payment Review", "Bi-Weekly", "Tuesday", "Vendor Payments", "USR-011", "Yes", "Yes", "High", "Pending", "", "Ensure MSME payments within 45 days"),
        ("ACT-024", "Expense Reimbursement Run", "Bi-Weekly", "Friday", "Vendor Payments", "USR-008", "Yes", "No", "Medium", "Pending", "", "Clear employee expenses"),

        # Compliance (MOD-005)
        ("ACT-025", "Income Tax Advance Tax", "Quarterly", "15", "Compliance", "USR-003", "Yes", "Yes", "High", "Pending", "https://incometax.gov.in", "Pay advance income tax installments"),
        ("ACT-026", "Income Tax Filing (Company)", "Yearly", "October", "Compliance", "USR-003", "Yes", "Yes", "High", "Pending", "", "Annual ITR filing"),
        ("ACT-027", "ROC Annual Return (MGT-7)", "Yearly", "November", "Compliance", "USR-015", "Yes", "Yes", "High", "Pending", "https://mca.gov.in", "File ROC returns"),
        ("ACT-028", "Financial Statements (AOC-4)", "Yearly", "October", "Compliance", "USR-015", "Yes", "Yes", "High", "Pending", "", "File financials with ROC"),
        ("ACT-029", "Secretarial Audit Report", "Yearly", "September", "Compliance", "USR-015", "Yes", "No", "Medium", "Pending", "", "Prepare secretarial audit"),
        ("ACT-030", "FEMA Compliance (FLA Return)", "Yearly", "July", "Compliance", "USR-004", "Yes", "No", "Medium", "Pending", "", "Foreign Liabilities and Assets return"),
        ("ACT-031", "Board Meeting Preparation", "Quarterly", "10", "Compliance", "USR-013", "Yes", "No", "High", "Pending", "", "Draft agenda and minutes"),

        # Audits (MOD-006)
        ("ACT-032", "Statutory Audit Prep", "Yearly", "May", "Audits", "USR-007", "Yes", "Yes", "High", "Pending", "", "Prepare schedules for statutory auditors"),
        ("ACT-033", "Tax Audit Prep", "Yearly", "August", "Audits", "USR-009", "Yes", "Yes", "High", "Pending", "", "Prepare Form 3CD schedules"),
        ("ACT-034", "Internal Audit (Q1)", "Yearly", "July", "Audits", "USR-014", "Yes", "No", "Medium", "Completed", "", "Conduct Q1 internal audit"),
        ("ACT-035", "Internal Audit (Q2)", "Yearly", "October", "Audits", "USR-014", "Yes", "No", "Medium", "Pending", "", "Conduct Q2 internal audit"),
        ("ACT-036", "Internal Audit (Q3)", "Yearly", "January", "Audits", "USR-014", "Yes", "No", "Medium", "Pending", "", "Conduct Q3 internal audit"),
        ("ACT-037", "Internal Audit (Q4)", "Yearly", "April", "Audits", "USR-014", "Yes", "No", "Medium", "Pending", "", "Conduct Q4 internal audit"),
        ("ACT-038", "Inventory Verification", "Yearly", "March", "Audits", "USR-009", "Yes", "Yes", "High", "Pending", "", "Physical stock verification"),

        # Financial Reports (MOD-007)
        ("ACT-039", "Daily Bank Reconciliation", "Daily", "", "Financial Reports", "USR-010", "No", "No", "Medium", "Pending", "", "Match daily bank statements"),
        ("ACT-040", "Monthly MIS Report", "Monthly", "5", "Financial Reports", "USR-011", "Yes", "Yes", "High", "Pending", "", "Publish monthly financials to management"),
        ("ACT-041", "Cash Flow Forecast", "Weekly", "Monday", "Financial Reports", "USR-004", "Yes", "No", "High", "Pending", "", "Update 13-week cash flow"),
        ("ACT-042", "Debtors Ageing Review", "Weekly", "Tuesday", "Financial Reports", "USR-001", "Yes", "No", "Medium", "Pending", "", "Follow up on overdue invoices"),
        ("ACT-043", "Fixed Asset Register Update", "Monthly", "28", "Financial Reports", "USR-012", "No", "No", "Low", "Pending", "", "Capitalize new assets and run depreciation"),
        ("ACT-044", "Intercompany Reconciliation", "Monthly", "12", "Financial Reports", "USR-008", "Yes", "No", "Medium", "Pending", "", "Match balances with subsidiary"),
        ("ACT-045", "Budget vs Actuals Variance", "Monthly", "8", "Financial Reports", "USR-013", "Yes", "Yes", "High", "Pending", "", "Analyze departmental variances"),
        
        # General / Admin
        ("ACT-046", "Backup Financial Data", "Weekly", "Friday", "Compliance", "USR-003", "Yes", "No", "High", "Pending", "", "Ensure ERP backups are secure"),
        ("ACT-047", "Review Software Subscriptions", "Monthly", "15", "Vendor Payments", "USR-011", "No", "No", "Low", "Pending", "", "Cancel unused SaaS tools"),
        ("ACT-048", "GST Clarification Replies", "Bi-Weekly", "Monday", "GST", "USR-001", "Yes", "No", "Medium", "Pending", "", "Check portal for notices"),
        ("ACT-049", "Employee Expense Audit", "Monthly", "25", "Audits", "USR-007", "Yes", "No", "Low", "Pending", "", "Sample check travel expenses"),
        ("ACT-050", "Insurance Policy Renewal", "Yearly", "March", "Compliance", "USR-015", "Yes", "Yes", "High", "Pending", "", "Renew D&O and Fire policies")
    ]
    
    for row in activities_list:
        ws_activities.append(row)
        
    # Data Validation for Activities
    dv_freq = DataValidation(type="list", formula1='"Daily,Weekly,Bi-Weekly,Monthly,Quarterly,Yearly"', allow_blank=False)
    ws_activities.add_data_validation(dv_freq)
    dv_freq.add("C2:C100")
    
    dv_status = DataValidation(type="list", formula1='"Pending,In Progress,Completed,Blocked"', allow_blank=True)
    ws_activities.add_data_validation(dv_status)
    dv_status.add("J2:J100")

    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws_activities.add_data_validation(dv_yesno)
    dv_yesno.add("G2:H100")

    dv_priority = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=False)
    ws_activities.add_data_validation(dv_priority)
    dv_priority.add("I2:I100")

    # -------------------------------------------------------------------------
    # 4. Reminder_Config Sheet
    # -------------------------------------------------------------------------
    ws_config = wb.create_sheet(title="Reminder_Config")
    ws_config.append(["Config_Key", "Config_Value"])
    ws_config.append(["Default_Dispatch_Time", "08:00 AM"])
    ws_config.append(["Max_Retries", "3"])
    ws_config.append(["Retry_Interval_Hours", "1"])
    ws_config.append(["Timezone", "Asia/Kolkata"])

    # -------------------------------------------------------------------------
    # 5. Email_Settings Sheet
    # -------------------------------------------------------------------------
    ws_email = wb.create_sheet(title="Email_Settings")
    ws_email.append(["Setting", "Value"])
    ws_email.append(["SMTP_Server", "smtp.gmail.com"])
    ws_email.append(["SMTP_Port", "587"])
    ws_email.append(["Sender_Email", "finance-reminders@hitech.local"])
    ws_email.append(["Enable_SSL", "True"])
    ws_email.append(["CC_Management", "cfo@example.com"])

    # -------------------------------------------------------------------------
    # 6. WhatsApp_Settings Sheet
    # -------------------------------------------------------------------------
    ws_wa = wb.create_sheet(title="WhatsApp_Settings")
    ws_wa.append(["Setting", "Value"])
    ws_wa.append(["Template_Name", "finance_task_reminder_v1"])
    ws_wa.append(["Language_Code", "en_US"])
    ws_wa.append(["Phone_Number_ID", "101234567890123"])
    ws_wa.append(["Business_Account_ID", "109876543210987"])

    # -------------------------------------------------------------------------
    # 7. Reminder_History Sheet (Historical Sent Records)
    # -------------------------------------------------------------------------
    ws_history = wb.create_sheet(title="Reminder_History")
    ws_history.append(["Run_ID", "Run_Date", "Activities_Sent", "Email_Status", "WhatsApp_Status", "Errors"])
    
    # Generate past 10 days of history
    base_date = datetime.now() - timedelta(days=10)
    for i in range(1, 11):
        run_date_str = (base_date + timedelta(days=i)).strftime("%Y-%m-%d")
        act_sent = random.randint(3, 12)
        email_stat = "Sent" if random.random() > 0.05 else "Failed"
        wa_stat = "Sent" if random.random() > 0.1 else "Failed"
        errors = ""
        if email_stat == "Failed":
            errors += "SMTP Timeout. "
        if wa_stat == "Failed":
            errors += "Graph API 500."
        
        ws_history.append([f"RUN-90{i:02d}", run_date_str, act_sent, email_stat, wa_stat, errors.strip()])

    # -------------------------------------------------------------------------
    # 8. Import_Logs Sheet (Previous Imports)
    # -------------------------------------------------------------------------
    ws_logs = wb.create_sheet(title="Import_Logs")
    ws_logs.append(["Import_ID", "Timestamp", "Sheet_Count", "Row_Count", "Status", "User_ID"])
    
    base_time = datetime.now() - timedelta(days=5)
    for i in range(1, 6):
        import_time = (base_time + timedelta(days=i, hours=random.randint(-4, 4))).strftime("%Y-%m-%d %H:%M:%S")
        ws_logs.append([f"IMP-20{i:02d}", import_time, "8", "150", "Success", f"USR-00{random.randint(1,4)}"])

    # -------------------------------------------------------------------------
    # Formatting
    # -------------------------------------------------------------------------
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Auto-adjust column widths
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = min(adjusted_width, 50) # Cap at 50

    wb.save(output_path)
    print(f"Workbook successfully updated with sample data at {output_path}")

if __name__ == "__main__":
    output = Path("data/Production_Reminders.xlsx")
    output.parent.mkdir(exist_ok=True)
    create_production_workbook(output)
